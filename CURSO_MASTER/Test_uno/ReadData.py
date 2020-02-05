import openpyxl

data=openpyxl.load_workbook("LoginData3.xlsx")

def numero_de_filas(Hoja):
    sh=data[Hoja]
    return  sh.max_row

def cell_data(Hoja, row, cell):
    sh=data[Hoja]
    cell=sh.cell(int(row),int(cell))
    return cell.value
    

print(numero_de_filas("Hoja1"))
print(cell_data("Hoja1",1,1))

